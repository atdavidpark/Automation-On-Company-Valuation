import os
import re
import time
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import xlsxwriter
import numpy as np
import pandas as pd

ROUNDING_DIGIT = 4
IS = "Income Statement"
BS = "Balance Sheet"
CF = "Cashflow Statement"

"""
IMPLEMENTATION:
    Get name of company from dataset.
    Change SUM algorithm.
    Fiscal year end.
    
FIXME BUGS:
    Empty row with same label disturbs excel_cell.
"""
def empty_unmodified(df, yrs_to_predict):
    unmodified = df.iloc[:, -yrs_to_predict] == '0'
    df.loc[unmodified, :] = np.nan
    df.index = [i if i not in list(df.index[unmodified]) else np.nan for i in list(df.index)]


def initialize_ratio_row(df, top_label, bot_label, new_label, nearby_label=None):
    """Create a new label and set a fractional formula for initialization."""
    df.loc[new_label] = [
        '={}/{}'.format(excel_cell(df, top_label, col, nearby_label),
                        excel_cell(df, bot_label, col))
        for col in df.columns
    ]


def insert_before(df, new_df, label):
    """Insert new DataFrame before the corresponding label row."""
    index = list(df.index).index(searched_label(df.index, label))
    return pd.concat([df.iloc[:index], new_df, df[index:]])


def insert_after(df, new_df, label):
    """Insert new DataFrame after the corresponding label row."""
    index = list(df.index).index(searched_label(df.index, label))
    return pd.concat([df.iloc[:index + 1], new_df, df[index + 1:]])


def add_empty_row(df):
    """Adds an empty row to the bottom of DataFrame."""
    df.loc["null"] = np.nan
    df.index = list(df.index)[:-1] + [np.nan]


def add_yr_column(df):
    """Appends one empty column representing year into DataFrame."""
    cur_yr = str(df.columns[len(df.columns) - 1])
    if cur_yr[-1] == 'E':
        cur_yr = str(int(cur_yr[:-1]) + 1) + 'E'
    else:
        cur_yr = str(int(cur_yr) + 1) + 'E'
    array = ['0' if i else np.nan for i in df.iloc[:,-1].notna().values]
    df.insert(len(df.columns), cur_yr, array)


def add_growth_rate_row(df, label, new_label):
    df.loc[new_label] = [np.nan] + [
        '={}/{}-1'.format(
            excel_cell(df, label, df.columns[i + 1]), excel_cell(df, label, df.columns[i])
        ) for i in range(len(df.columns) - 1)
    ]


def driver_extend(df, row_label, how, last_given_yr, yrs_to_predict, num_excluded=0):
    formula = ""
    if how is "round":
        formula = "=ROUND(" + excel_cell(df, row_label, last_given_yr) + ',' + \
                  str(ROUNDING_DIGIT) + ')'
    elif how is "avg":
        formula = "=AVERAGE(" + excel_cell(df, row_label, df.columns[0 + num_excluded]) + ':' + \
                  excel_cell(df, row_label, last_given_yr) + ')'
    df.loc[row_label].iloc[-yrs_to_predict] = formula
    temp = excel_cell(df, row_label, df.columns[-yrs_to_predict])
    df.loc[row_label].iloc[-yrs_to_predict + 1:] = '=' + temp


def fixed_extend(df, row_label, how, yrs):
    """Predicts the corresponding row of data only using data from current row."""
    if how is "prev":
        df.at[row_label, df.columns[-yrs:]] = df.loc[row_label, df.columns[-yrs - 1]]
    elif how is "avg":
        mean = df.loc[row_label].iloc[:-yrs].mean(axis=0)
        df.at[row_label, df.columns[-yrs]] = mean
    elif how is "mix":
        mean = df.loc[row_label].iloc[:-3].mean(axis=0)
        if abs(mean - df.loc[row_label, df.columns[-2]]) > mean * 0.5:
            df.at[row_label, df.columns[-1]] = df.loc[row_label].iloc[:-1].mean(axis=0)
        else:
            df.at[row_label, df.columns[-1]] = df.loc[row_label, df.columns[-2]]
    elif how is "zero":
        df.at[row_label, df.columns[-yrs:]] = 0
    else:
        print("ERROR: fixed_extend")
        exit(1)


def eval_formula(df, formula):
    """Evaluates an excel formula of a dataframe.
    The mathematical operations must decrease in priority from left to right."""
    ans = 0
    cells = re.findall(r"[A-Z][0-9]*", formula)
    ops = ['+'] + re.findall(r"[+|-|*|/|]", formula)
    
    for i in range(len(cells)):
        row = int(cells[i][1:]) - 2
        col = ord(cells[i][0]) - ord('A') - 1
        if ops[i] is '+':
            ans += df.iat[row, col]
        elif ops[i] is '-':
            ans -= df.iat[row, col]
        elif ops[i] is '*':
            ans *= df.iat[row, col]
        elif ops[i] is '/':
            ans /= df.iat[row, col]
        else:
            print("ERROR: Invalid operator symbol")
            exit(1)
    return ans


def excel_cell(df, row_label, col_label, nearby_label=None):
    """Returns corresponding excel cell position given row label and column label. 
    Note that if there are more than 26 columns, this function does not work properly."""
    if not row_label:
        print("ERROR: excel_cell")
        exit(1)
    letter = chr(ord('A') + df.columns.get_loc(col_label) + 2)
    row_mask = df.index.get_loc(row_label)
    if type(row_mask) is int:
        return letter + str(3 + row_mask)
    else:
        nearby_index = df.index.get_loc(nearby_label)
        matched_indices = [i for i, j in enumerate(row_mask) if j]
        distance_vals = [abs(nearby_index - i) for i in matched_indices]
        return letter + str(3 + matched_indices[distance_vals.index(min(distance_vals))])


def searched_label(labels, target):
    """Returns target label from a list of DataFrame labels."""
    score_dict = {label: 0 for label in labels}

    for word in target.split():
        for label in set(labels):
            if word.lower() in str(label).lower():
                score_dict[label] += 1

    if sum(score_dict.values()) == 0:
        return ""
    def compare(pair):
        if type(pair[0]) is str:
            return len(pair[0])
        return 0
    return max(sorted(score_dict.items(), key=compare), key=lambda pair: pair[1])[0]


def filter_in(df, filters):
    labels = [searched_label(df.index, label) for label in filters] + \
             list(df[df.iloc[:, 1].isna()].index)
    df.drop([label for label in df.index if label not in labels], inplace=True)


def preprocess(df, filter_in=[]):
    # Reverse columns
    df = df.loc[:, ::-1]

    # Replace all '-' with 0
    df = df.replace('-', 0)

    # Delete current data
    if df.iat[0, -1] == 'LTM':
        df = df.iloc[:, :-1]

    # Remove the row with number of days
    df = df[1:]

    # Change dates to only years
    fye = df.columns[1]
    df.columns = [
        '20' + ''.join([char for char in column if char.isdigit()]) for column in df.columns
    ]
    return df, fye


def process_is(is_df, cf_df, growth_rates, yrs_to_predict):
    # Short-hands
    first_yr = is_df.columns[0]
    last_given_yr = is_df.columns[-1]

    # Insert 4 empty rows
    is_df = pd.concat(
        [pd.DataFrame({yr: [np.nan] * 4 for yr in is_df.columns}, index=[np.nan] * 4), is_df]
    )
    cf_df = pd.concat(
        [pd.DataFrame({yr: [np.nan] * 4 for yr in cf_df.columns}, index=[np.nan] * 4), cf_df]
    )

    # Income statement labels
    sales = searched_label(is_df.index, "total sales")
    cogs = searched_label(is_df.index, "cost of goods sold")
    is_df.index = [
        i if i != cogs else "Cost of Goods Sold (COGS) excl. D&A" for i in list(is_df.index)
    ]
    cogs = "Cost of Goods Sold (COGS) excl. D&A"
    gross_income = searched_label(is_df.index, "gross income")
    sgna = searched_label(is_df.index, "sg&a expense")
    other_expense = searched_label(is_df.index, "other expense")
    ebit = searched_label(is_df.index, "ebit operating income")
    nonoperating_income = searched_label(is_df.index, "nonoperating income net")
    interest_expense = searched_label(is_df.index, "interest expense")
    unusual = searched_label(is_df.index, "unusual expense")
    income_tax = searched_label(is_df.index, "income taxes")
    diluted_eps = searched_label(is_df.index, "dilute eps")
    net_income = searched_label(is_df.index, "net income")
    div_per_share = searched_label(is_df.index, "div per share")
    ebitda = searched_label(is_df.index, "ebitda")

    # Insert pretax income row before income taxes
    pretax_df = pd.DataFrame(
        {
            yr: '={}+{}-{}-{}'.format(
                excel_cell(is_df, ebit, yr), excel_cell(is_df, nonoperating_income, yr),
                excel_cell(is_df, interest_expense, yr), excel_cell(is_df, unusual, yr)
            ) for yr in is_df.columns
        }, index=["Pretax Income"]
    )
    is_df = insert_before(is_df, pretax_df, "income taxes")
    pretax = "Pretax Income"

    # Insert depreciation & amortization expense before SG&A expense
    dna_expense_df = pd.DataFrame(
        {
            yr: "='{}'!".format(CF) + excel_cell(
                cf_df, searched_label(cf_df.index, "deprecia & amortiza expense"), yr
            ) for yr in is_df.columns
        }, index=["Depreciation & Amortization Expense"]
    )
    is_df = insert_before(is_df, dna_expense_df, "sg&a expense")
    dna_expense = "Depreciation & Amortization Expense"

    # Insert diluted shares outstanding before dividends per share
    diluted_share_outstanding_df = pd.DataFrame(
        {
            yr: "={}/{}".format(
                excel_cell(is_df, net_income, yr), excel_cell(is_df, diluted_eps, yr)
            ) for yr in is_df.columns
        }, index=["Diluted Shares Outstanding"]
    )
    is_df = insert_before(is_df, diluted_share_outstanding_df, "div per share")
    diluted_share_outstanding = "Diluted Shares Outstanding"

    # Recalculate COGS
    is_df.loc[cogs] = [
        '={}-{}'.format(is_df.at[cogs, yr], excel_cell(is_df, dna_expense, yr))
        for yr in is_df.columns
    ]

    # Add driver rows to income statement
    add_empty_row(is_df)
    is_df.loc["Driver Ratios"] = np.nan
    add_growth_rate_row(is_df, sales, "Sales Growth %")
    sales_growth = "Sales Growth %"
    add_empty_row(is_df)
    initialize_ratio_row(is_df, cogs, sales, "COGS Sales Ratio")
    cogs_ratio = "COGS Sales Ratio"
    add_empty_row(is_df)
    initialize_ratio_row(is_df, sgna, sales, "SG&A Sales Ratio")
    sgna_ratio = "SG&A Sales Ratio"
    add_empty_row(is_df)
    initialize_ratio_row(is_df, dna_expense, sales, "D&A Sales Ratio")
    dna_ratio = "D&A Sales Ratio"
    add_empty_row(is_df)
    initialize_ratio_row(is_df, unusual, ebit, "Unusual Expense EBIT Ratio")
    unusual_ratio = "Unusual Expense EBIT Ratio"
    add_empty_row(is_df)
    initialize_ratio_row(is_df, income_tax, pretax, "Effective Tax Rate")
    effective_tax = "Effective Tax Rate"

    # Add prediction years
    for i in range(yrs_to_predict):
        add_yr_column(is_df)

    # Append growth rates to driver row
    is_df.loc[sales_growth].iloc[-yrs_to_predict:] = growth_rates

    # Calculate driver ratios
    initialize_ratio_row(is_df, div_per_share, diluted_eps, "Dividend Payout Ratio")
    initialize_ratio_row(is_df, ebitda, sales, "EBITDA Margin", sales_growth)
    is_df.loc[dna_ratio].iloc[-yrs_to_predict:] = is_df.loc[dna_ratio, last_given_yr]
    driver_extend(is_df, cogs_ratio, "round", last_given_yr, yrs_to_predict)
    driver_extend(is_df, sgna_ratio, "round", last_given_yr, yrs_to_predict)
    driver_extend(is_df, unusual_ratio, "avg", last_given_yr, yrs_to_predict)
    driver_extend(is_df, diluted_share_outstanding, "avg", last_given_yr, yrs_to_predict, 3)
    driver_extend(is_df, effective_tax, "avg", last_given_yr, yrs_to_predict)

    # Calculate fixed variables
    fixed_extend(is_df, nonoperating_income, 'prev', yrs_to_predict)
    fixed_extend(is_df, interest_expense, 'prev', yrs_to_predict)
    fixed_extend(is_df, other_expense, 'prev', yrs_to_predict)
    fixed_extend(is_df, div_per_share, 'prev', yrs_to_predict)

    # Calculate net income
    is_df.loc[net_income] = [
        '={}-{}'.format(excel_cell(is_df, pretax, yr), excel_cell(is_df, income_tax, yr))
        for yr in is_df.columns
    ]

    for i in range(yrs_to_predict):
        cur_yr = is_df.columns[-yrs_to_predict + i]
        prev_yr = is_df.columns[-yrs_to_predict + i - 1]    

        # Calculate variables
        is_df.at[sales, cur_yr] = '={}*(1+{})'.format(
            excel_cell(is_df, sales, prev_yr), excel_cell(is_df, sales_growth, cur_yr)
        )
        is_df.at[cogs, cur_yr] = '={}*{}'.format(
            excel_cell(is_df, sales, cur_yr), excel_cell(is_df, cogs_ratio, cur_yr)
        )
        is_df.at[gross_income, cur_yr] = '={}-{}'.format(
            excel_cell(is_df, sales, cur_yr), excel_cell(is_df, cogs, cur_yr)
        )
        is_df.at[dna_expense, cur_yr] = '={}*{}'.format(
            excel_cell(is_df, sales, cur_yr), excel_cell(is_df, dna_ratio, cur_yr)
        )
        is_df.at[sgna, cur_yr] = '={}*{}'.format(
            excel_cell(is_df, sales, cur_yr), excel_cell(is_df, sgna_ratio, cur_yr)
        )
        is_df.at[ebit, cur_yr] = '={}-{}-{}'.format(
            excel_cell(is_df, gross_income, cur_yr), excel_cell(is_df, sgna, cur_yr),
            excel_cell(is_df, other_expense, cur_yr)
        )
        is_df.at[unusual, cur_yr] = '={}*{}'.format(
            excel_cell(is_df, ebit, cur_yr), excel_cell(is_df, unusual_ratio, cur_yr)
        )
        is_df.at[pretax, cur_yr] = '={}+{}-{}-{}'.format(
            excel_cell(is_df, ebit, cur_yr), excel_cell(is_df, nonoperating_income, cur_yr),
            excel_cell(is_df, interest_expense, cur_yr), excel_cell(is_df, unusual, cur_yr)
        )
        is_df.at[income_tax, cur_yr] = '={}*{}'.format(
             excel_cell(is_df, pretax, cur_yr), excel_cell(is_df, effective_tax, cur_yr)
        )
        is_df.at[diluted_eps, cur_yr] = '={}/{}'.format(
            excel_cell(is_df, net_income, cur_yr),
            excel_cell(is_df, diluted_share_outstanding,cur_yr)
        )
        is_df.at[ebitda, cur_yr] = '={}+{}'.format(
            excel_cell(is_df, dna_expense, cur_yr), excel_cell(is_df, ebit, cur_yr)
        )
    empty_unmodified(is_df, yrs_to_predict)

    return is_df, cf_df


def process_bs(is_df, bs_df, cf_df, yrs_to_predict):
    # Short-hands
    first_yr = bs_df.columns[0]
    last_given_yr = bs_df.columns[-1]

    # Insert 4 empty rows
    bs_df = pd.concat(
        [pd.DataFrame({yr: [np.nan] * 4 for yr in bs_df.columns}, index=[np.nan] * 4), bs_df]
    )

    # Balance sheet labels
    st_receivables = searched_label(bs_df.index, "short term receivable")
    cash_st_investments = searched_label(bs_df.index, "cash short term investment")
    inventories = searched_label(bs_df.index, "inventor")
    other_cur_assets = searched_label(bs_df.index, "other current asset")
    total_cur_assets = searched_label(bs_df.index, "total current asset")
    net_property_plant_equipment = searched_label(bs_df.index, "net property plant qquipment")
    total_investments_n_advances = searched_label(bs_df.index, "total investment advance")
    intangible_assets = searched_label(bs_df.index, "intangible asset")
    deferred_tax_assets = searched_label(bs_df.index, "deferred tax asset")
    other_assets = searched_label(bs_df.index, "other asset")
    total_assets = searched_label(bs_df.index, "total asset")
    st_debt_n_cur_portion_lt_debt = searched_label(bs_df.index, "debt st lt cur portion")
    accounts_payable = searched_label(bs_df.index, "account payable")
    income_tax_payable = searched_label(bs_df.index, "income tax payable")
    other_cur_liabilities = searched_label(bs_df.index, "other current liabilities")
    total_cur_liabilities = searched_label(bs_df.index, "total current liabilities")
    lt_debt = searched_label(bs_df.index, "long term debt")
    provision_for_risks_n_charges = searched_label(bs_df.index, "provision for risk & charge")
    deferred_tax_liabilities = searched_label(bs_df.index, "deferred tax liabilities")
    other_liabilities = searched_label(bs_df.index, "other liabilities")
    total_liabilities = searched_label(bs_df.index, "total liabilities")
    total_shareholder_equity = searched_label(bs_df.index, "total shareholder equity")
    total_liabilities_n_shareholders_equity = searched_label(
        bs_df.index, "total liabilities shareholder equity"
    )

    # Income statement labels
    sales = searched_label(is_df.index, "total sales")
    cogs = searched_label(is_df.index, "cost of goods sold")
    net_income = searched_label(is_df.index, "net income")

    # Cash flow statement labels
    deprec_deplet_n_amort = searched_label(cf_df.index, "depreciation depletion amortization") 
    capital_expenditures = searched_label(cf_df.index, "capital expenditure")
    cash_div_paid = searched_label(cf_df.index, "cash div paid")
    change_in_capital_stock = searched_label(cf_df.index, "change in capital stock")

    # Add driver rows to balance sheet
    add_empty_row(bs_df)
    bs_df.loc["Driver Ratios"] = np.nan
    # DSO
    bs_df.loc["DSO"] = [
        "={}/'{}'!{}*365".format(
            excel_cell(bs_df, st_receivables, yr), IS, excel_cell(is_df, sales, yr)
        ) for yr in bs_df.columns
    ]
    dso = "DSO"
    # Other current assets growth %
    add_growth_rate_row(bs_df, other_cur_assets, "Other Current Assets Growth %")
    other_cur_assets_growth = "Other Current Assets Growth %"
    # DPO
    add_empty_row(bs_df)
    bs_df.loc["DPO"] = [
        "={}/'{}'!{}*366".format(
            excel_cell(bs_df, accounts_payable, yr), IS, excel_cell(is_df, cogs, yr)
        ) for yr in bs_df.columns
    ]
    dpo = "DPO"
    # Miscellaneous Current Liabilities Growth %
    add_growth_rate_row(bs_df, other_cur_liabilities, "Miscellaneous Current Liabilities Growth %")
    misc_cur_liabilities_growth = "Miscellaneous Current Liabilities Growth %"
    # Inventory turnober ratio
    add_empty_row(bs_df)
    bs_df.loc["Inventory Turnover Ratio"] = np.nan
    bs_df.loc["Inventory Turnover Ratio"].iloc[1:] = [
        "='{}'!{}/({}+{})*2".format(
            IS, excel_cell(is_df, cogs, bs_df.columns[i + 1]),
            excel_cell(bs_df, inventories, bs_df.columns[i]),
            excel_cell(bs_df, inventories, bs_df.columns[i+1])
        ) for i in range(len(bs_df.columns) - 1)
    ]
    inventory_ratio = "Inventory Turnover Ratio"

    # Add driver rows to cash flow statement
    add_empty_row(cf_df)
    cf_df.loc["Driver Ratios"] = np.nan
    # Capital Expenditure Revenue Ratio
    cf_df.loc["Capital Expenditure Revenue Ratio"]  = [
        "=-{}/'{}'!{}".format(
            excel_cell(cf_df, capital_expenditures, yr), IS, excel_cell(is_df, sales, yr)
        ) for yr in cf_df.columns
    ]
    # Other Funds Net Operating CF Ratio
    net_operating_cf = searched_label(cf_df.index, "net operat cash flow cf")
    initialize_ratio_row(cf_df, searched_label(cf_df.index, "other funds"), net_operating_cf,
                         "Other Funds Net Operating CF Ratio", net_operating_cf)

    # Add prediction years
    for i in range(yrs_to_predict):
        add_yr_column(bs_df)
    for i in range(yrs_to_predict):
        add_yr_column(cf_df)

    # Insert cash balance
    cash_balance_df = pd.DataFrame({yr: np.nan for yr in cf_df.columns}, index=["Cash Balance"])
    cf_df = insert_after(cf_df, cash_balance_df, "net change in tax")
    cash_balance = searched_label(cf_df.index, "cash balance")

    # Inesrt working capital row
    wk_df = pd.DataFrame(
        {
            yr: ['={}-{}'.format(
                excel_cell(bs_df, total_cur_assets, yr), 
                excel_cell(bs_df, total_cur_liabilities, yr)
            ), np.nan] for yr in bs_df.columns
        }, index=["Working Capital", np.nan]
    )
    bs_df = insert_before(bs_df, wk_df, "driver ratios")

    # Inesrt balance row
    balance_df = pd.DataFrame(
        {
            yr: ['={}-{}'.format(
                excel_cell(bs_df, total_assets, yr), 
                excel_cell(bs_df, total_liabilities_n_shareholders_equity, yr)
            ), np.nan] for yr in bs_df.columns
        }, index=["Balance", np.nan]
    )
    bs_df = insert_before(bs_df, balance_df, "working capital")

    # Calculate driver ratios
    bs_df.loc[dso].iloc[-yrs_to_predict:] = '=' + excel_cell(
        bs_df,dso, bs_df.columns[-yrs_to_predict - 2]
    )
    driver_extend(bs_df, dpo, "avg", last_given_yr, yrs_to_predict)
    driver_extend(bs_df, other_cur_assets_growth, "avg", last_given_yr, yrs_to_predict)
    driver_extend(bs_df, misc_cur_liabilities_growth, "avg", last_given_yr, yrs_to_predict)
    driver_extend(bs_df, inventory_ratio, "avg", last_given_yr, yrs_to_predict)

    # Calculate fixed variables
    fixed_extend(bs_df, inventories, 'prev', yrs_to_predict)
    fixed_extend(bs_df, total_investments_n_advances, 'prev', yrs_to_predict)
    fixed_extend(bs_df, intangible_assets, 'prev', yrs_to_predict)
    fixed_extend(bs_df, deferred_tax_assets, 'prev', yrs_to_predict)
    fixed_extend(bs_df, other_assets, 'prev', yrs_to_predict)
    fixed_extend(bs_df, st_debt_n_cur_portion_lt_debt, 'prev', yrs_to_predict)
    fixed_extend(bs_df, income_tax_payable, 'prev', yrs_to_predict)
    fixed_extend(bs_df, lt_debt, 'prev', yrs_to_predict)
    fixed_extend(bs_df, provision_for_risks_n_charges, 'prev', yrs_to_predict)
    fixed_extend(bs_df, deferred_tax_liabilities, 'prev', yrs_to_predict)
    fixed_extend(bs_df, other_liabilities, 'prev', yrs_to_predict)

    # Calculate total liabilities & shareholders' equity
    bs_df.loc[total_liabilities_n_shareholders_equity] = [
        '={}+{}'.format(
            excel_cell(bs_df, total_liabilities, yr),
            excel_cell(bs_df, total_shareholder_equity, yr)
        ) for yr in is_df.columns
    ]

    for i in range(yrs_to_predict):
        cur_yr = bs_df.columns[-yrs_to_predict + i]
        prev_yr = bs_df.columns[-yrs_to_predict + i - 1] 

        # Calculate variables
        bs_df.at[cash_st_investments, cur_yr] = "='{}'!{}".format(
            CF, excel_cell(cf_df, cash_balance, cur_yr)
        )
        bs_df.at[st_receivables, cur_yr] = "={}/365*'{}'!{}".format(
            excel_cell(bs_df, dso, cur_yr), IS, excel_cell(is_df, sales, cur_yr)
        )
        bs_df.at[other_cur_assets, cur_yr] = '={}*(1+{})'.format(
            excel_cell(bs_df, other_cur_assets, prev_yr),
            excel_cell(bs_df, other_cur_assets_growth, cur_yr)
        )
        bs_df.at[net_property_plant_equipment, cur_yr] = "={}-'{}'!{}-'{}'!{}".format(
            excel_cell(bs_df, net_property_plant_equipment, prev_yr), CF,
            excel_cell(cf_df, deprec_deplet_n_amort, cur_yr), CF,
            excel_cell(cf_df, capital_expenditures, cur_yr)
        )
        bs_df.at[accounts_payable, cur_yr] = "={}/365*'{}'!{}".format(
            excel_cell(bs_df, dpo, cur_yr), IS, excel_cell(is_df, cogs, cur_yr)
        )
        bs_df.at[other_cur_liabilities, cur_yr] = "={}*(1+{})".format(
            excel_cell(bs_df, other_cur_liabilities, prev_yr),
            excel_cell(bs_df, misc_cur_liabilities_growth, cur_yr)
        )

        # FIXME sum positions may not be correct
        bs_df.at[total_cur_assets, cur_yr] = '=SUM({}:{})'.format(
            excel_cell(bs_df, cash_st_investments, cur_yr),
            excel_cell(bs_df, other_cur_assets, cur_yr)
        )
        bs_df.at[total_assets, cur_yr] = '={}+SUM({}:{})'.format(
            excel_cell(bs_df, total_cur_assets, cur_yr),
            excel_cell(bs_df, net_property_plant_equipment, cur_yr),
            excel_cell(bs_df, other_assets, cur_yr)
        )
        bs_df.at[total_cur_liabilities, cur_yr] = '=SUM({}:{})'.format(
            excel_cell(bs_df, st_debt_n_cur_portion_lt_debt, cur_yr),
            excel_cell(bs_df, other_cur_liabilities, cur_yr)
        )
        bs_df.at[total_liabilities, cur_yr] = '={}+SUM({}:{})'.format(
            excel_cell(bs_df, total_cur_liabilities, cur_yr),
            excel_cell(bs_df, lt_debt, cur_yr), excel_cell(bs_df, other_liabilities, cur_yr)
        )
        bs_df.at[total_shareholder_equity, cur_yr] = "={}+'{}'!{}+'{}'!{}+'{}'!{}".format(
            excel_cell(bs_df, total_shareholder_equity, prev_yr), CF,
            excel_cell(cf_df, change_in_capital_stock, cur_yr), IS,
            excel_cell(is_df, net_income, cur_yr), CF, excel_cell(cf_df, cash_div_paid, cur_yr)
        )

    empty_unmodified(bs_df, yrs_to_predict)

    return bs_df, cf_df


def process_cf(is_df, bs_df, cf_df, yrs_to_predict):
    # Short-hands
    first_yr = cf_df.columns[0]
    last_given_yr = cf_df.columns[-yrs_to_predict-1]

    # Cash flow statement labels
    net_income_cf = searched_label(cf_df.index, "net income")
    deprec_deplet_n_amort = searched_label(cf_df.index, "depreciation depletion amortization")
    deferred_taxes = searched_label(cf_df.index, "deferred taxes")
    other_funds = searched_label(cf_df.index, "other funds")
    funds_from_operations = searched_label(cf_df.index, "fund from operation")
    changes_in_working_capital = searched_label(cf_df.index, "change work capital")
    net_operating_cf = searched_label(cf_df.index, "net operat cash flow cf")
    capital_expenditures = searched_label(cf_df.index, "capital expenditure")
    net_asset_acquisition = searched_label(cf_df.index, "net asset acquisiton")
    fixed_assets_n_businesses_sale = searched_label(cf_df.index, "fixed asset sale business")
    purchase_sale_of_investments = searched_label(cf_df.index, "purchase sale investment")
    net_investing_cf = searched_label(cf_df.index, "net invest cash flow")
    cash_div_paid = searched_label(cf_df.index, "cash div paid")
    change_in_capital_stock = searched_label(cf_df.index, "change in capital stock")
    net_inssuance_reduction_of_debt = searched_label(cf_df.index, "net issuance reduct debt")
    net_financing_cf = searched_label(cf_df.index, "net financ cash flow cf")
    net_change_in_cash = searched_label(cf_df.index, "net change in cash")
    cash_balance = searched_label(cf_df.index, "cash balance")
    capex_ratio = "Capital Expenditure Revenue Ratio"
    other_funds_net_operating_ratio = "Other Funds Net Operating CF Ratio"

    # Income statement labels
    sales = searched_label(is_df.index, "total sales")
    deprec_amort_expense = searched_label(is_df.index, "depreciation amortization expense")
    net_income_is = searched_label(is_df.index, "net income")
    diluted_share_outstanding = searched_label(is_df.index, "diluted share outstanding")
    div_per_share = searched_label(is_df.index, "dividend per share")

    # Balance sheet labels
    other_cur_assets = searched_label(bs_df.index, "other current asset")
    other_cur_liabilities = searched_label(bs_df.index, "other current liabilit")
    cash_st_investments = searched_label(bs_df.index, "cash short term investment")
    st_receivables = searched_label(bs_df.index, "short term receivable")
    total_cur_assets = searched_label(bs_df.index, "total current asset")
    st_debt_n_cur_portion_lt_debt = searched_label(bs_df.index, "st debt cur portion lt")
    accounts_payable = searched_label(bs_df.index, "account payable")
    total_cur_liabilities = searched_label(bs_df.index, "total current liabilit")
    lt_debt = searched_label(bs_df.index, "long term debt")

    # Insert cash balance
    cf_df.loc["Cash Balance"].iloc[-yrs_to_predict - 1:] = [
        "='{}'!{}".format(BS, excel_cell(bs_df, cash_st_investments, last_given_yr))
    ] + [
        '={}+{}'.format(
            excel_cell(cf_df, "Cash Balance", cf_df.columns[-yrs_to_predict + i - 1]),
            excel_cell(cf_df, net_change_in_cash, cf_df.columns[-yrs_to_predict + i])
        ) for i in range(yrs_to_predict)
    ]

    # Add levered free CF row
    cf_df.loc["Levered Free Cash Flow"] = [
        '={}+{}'.format(
            excel_cell(cf_df, net_operating_cf, yr),
            excel_cell(cf_df, capital_expenditures, yr)
        ) for yr in cf_df.columns
    ]
    levered_free_cf = "Levered Free Cash Flow"

    # Add levered free CF row growth %
    add_growth_rate_row(cf_df, levered_free_cf, "Levered Free Cash Flow Growth %")
    levered_free_cf_growth = "Levered Free Cash Flow Growth %"

    # Calculate driver ratios
    driver_extend(cf_df, capex_ratio, "avg", last_given_yr, yrs_to_predict)
    driver_extend(cf_df, other_funds_net_operating_ratio, "avg", last_given_yr, yrs_to_predict)

    # Calculate fixed variables
    fixed_extend(cf_df, deferred_taxes, "zero", yrs_to_predict)
    fixed_extend(cf_df, other_funds, "zero", yrs_to_predict)
    fixed_extend(cf_df, net_asset_acquisition, "zero", yrs_to_predict)
    fixed_extend(cf_df, fixed_assets_n_businesses_sale, "zero", yrs_to_predict)
    fixed_extend(cf_df, purchase_sale_of_investments, "zero", yrs_to_predict)
    fixed_extend(cf_df, change_in_capital_stock, "prev", yrs_to_predict)

    # Calculate net operating CF
    cf_df.loc[net_operating_cf] = [
        '={}+{}'.format(
            excel_cell(cf_df, funds_from_operations, yr),
            excel_cell(cf_df, changes_in_working_capital, yr)
        ) for yr in cf_df.columns
    ]

    # Calculate net investing CF
    cf_df.loc[net_investing_cf] = [
        '=SUM({}:{})'.format(
            excel_cell(cf_df, capital_expenditures, yr),
            excel_cell(cf_df, other_funds, yr, nearby_label=net_investing_cf)
        ) for yr in cf_df.columns
    ]

    # Calcualate net financing CF
    cf_df.loc[net_financing_cf] = [
        '=SUM({}:{})'.format(
            excel_cell(cf_df, cash_div_paid, yr),
            excel_cell(cf_df, other_funds, yr, nearby_label=net_financing_cf)
        ) for yr in cf_df.columns
    ]

    # Calculate net change in cash
    cf_df.loc[net_change_in_cash] = [
        '={}+{}+{}'.format(
            excel_cell(cf_df, net_operating_cf, yr), excel_cell(cf_df, net_investing_cf, yr),
            excel_cell(cf_df, net_financing_cf, yr)
        ) for yr in cf_df.columns
    ]

    for i in range(yrs_to_predict):
        cur_yr = is_df.columns[-yrs_to_predict + i]
        prev_yr = is_df.columns[-yrs_to_predict + i - 1]    

        # Calculate variables
        cf_df.at[net_income_cf, cur_yr] = "='{}'!{}".format(
            IS, excel_cell(is_df, net_income_is, cur_yr)
        )
        cf_df.at[deprec_deplet_n_amort, cur_yr] = "='{}'!{}".format(
            IS, excel_cell(is_df, deprec_amort_expense, cur_yr)
        )

        cf_df.at[funds_from_operations, cur_yr] = "=SUM({}:{})".format(
            excel_cell(cf_df, net_income_cf, cur_yr),
            excel_cell(cf_df, other_funds, cur_yr, nearby_label=net_operating_cf)
        )

        cf_df.at[changes_in_working_capital, cur_yr] = "=SUM('{}'!{}:{})-SUM('{}'!{}:{})".format(
            BS, excel_cell(bs_df, st_receivables, prev_yr),
            excel_cell(bs_df, other_cur_assets, prev_yr),
            BS, excel_cell(bs_df, st_receivables, cur_yr),
            excel_cell(bs_df, other_cur_assets, cur_yr)
        )
        cf_df.at[changes_in_working_capital, cur_yr] += "+SUM('{}'!{}:{})-SUM('{}'!{}:{})".format(
            BS, excel_cell(bs_df, accounts_payable, cur_yr),
            excel_cell(bs_df, other_cur_liabilities, cur_yr),
            BS, excel_cell(bs_df, accounts_payable, prev_yr),
            excel_cell(bs_df, other_cur_liabilities, prev_yr)
        )

        cf_df.at[capital_expenditures, cur_yr] = "=-'{}'!{}*{}".format(
            IS, excel_cell(is_df, sales, cur_yr), excel_cell(cf_df, capex_ratio, cur_yr)
        )
        cf_df.at[cash_div_paid, cur_yr] = "=-'{}'!{}*'{}'!{}".format(
            IS, excel_cell(is_df, diluted_share_outstanding, cur_yr),
            IS, excel_cell(is_df, div_per_share, cur_yr)
        )
        cf_df.at[net_inssuance_reduction_of_debt, cur_yr] = "='{}'!{}-'{}'!{}".format(
            BS, excel_cell(bs_df, lt_debt, cur_yr),
            BS, excel_cell(bs_df, lt_debt, prev_yr)
        )
    empty_unmodified(cf_df, yrs_to_predict)

    return cf_df


def style_range(ws, start, end, fill=PatternFill(), font=Font(), border=Border(),
                alignment=Alignment()):
    letter1, num1 = start[0], start[1:]
    letter2, num2 = end[0], end[1:]
    if letter1 == letter2:  # column
        for i in range(int(num1), int(num2) + 1):
            ws[letter1 + str(i)].font = font
            ws[letter1 + str(i)].fill = fill
            ws[letter1 + str(i)].border = border
            ws[letter1 + str(i)].alignment = alignment
    elif num1 == num2:  # row
        for i in range(ord(letter2) - ord(letter1) + 1):
            ws[chr(ord(letter1) + i) + num1].font = font
            ws[chr(ord(letter1) + i) + num1].fill = fill
            ws[chr(ord(letter1) + i) + num1].border = border
            ws[chr(ord(letter1) + i) + num1].alignment = alignment
    else:
        print("ERROR: style_range")
        exit(1)


def style_ws(ws, sheet_name, is_df, bs_df, cf_df, fye):
    border = Side(border_style="thin", color="000000")

    # Insert empty column to beginning
    ws.insert_cols(1)

    letter, num = ws.dimensions.split(':')[1][0], ws.dimensions.split(':')[1][1:]

    ws.sheet_view.showGridLines = False  # No grid lines
    ws.move_range('C1:' + letter + '1', rows=4)  # Move year row down
    ws.column_dimensions['B'].width = 50  # Change width of labels

    ws['B2'] = sheet_name
    ws['B2'].font = Font(bold=True)
    ws['B2'].fill = PatternFill(fill_type='solid', fgColor='bababa')
    ws['B3'] = "($ in millions of U.S. Dollar)"
    ws['B3'].font = Font(italic=True)
    style_range(ws, 'B3', letter + '3', fill=PatternFill(fill_type='solid', fgColor='bababa'))

    # Central element Annual
    ws[chr((ord('C') + ord(letter)) // 2) + '3'] = "Annual"
    ws[chr((ord('C') + ord(letter)) // 2) + '3'].font = Font(bold=True)
    ws[chr((ord('C') + ord(letter)) // 2) + '4'] = "FYE " + fye

    # Year row style
    style_range(ws, 'C5', letter + '5', font=Font(bold=True, underline="single"),
                border=Border(top=border, bottom=border),
                alignment=Alignment(horizontal="center", vertical="center"))

    # Label column
    style_range(ws, 'B7', 'B' + num, fill=PatternFill(fill_type='solid', fgColor='dddddd'))

    # Style sum rows
    for cell in [letter + str(i + 1) for i in range(int(num) - 1)]:
        if type(ws[cell].value) is str and 'SUM' in ws[cell].value and len(ws[cell].value) < 30:
            num = cell[1:]
            ws['B' + num].font = Font(bold=True)
            style_range(ws, 'C' + num, letter + num, font=Font(bold=True),
                        border=Border(top=border))

    # Style specific rows
    def style_row(ws, label, sheet_name, is_df, bs_df, cf_df, border_bool=True, bold_bool=True,
                  underline=None):
        df = None
        num = 0
        if sheet_name == IS:
            num = str(int(excel_cell(is_df, searched_label(is_df.index, label),
                                     is_df.columns[0])[1:]))
        elif sheet_name == BS:
            num = str(int(excel_cell(bs_df, searched_label(bs_df.index, label),
                                     bs_df.columns[0])[1:]))
        elif sheet_name == CF:
            num = str(int(excel_cell(cf_df, searched_label(cf_df.index, label),
                                     cf_df.columns[0])[1:]))
        ws['B' + num].font = Font(bold=True, underline=underline)
        border_style = Border(top=border) if border_bool else Border()
        bold_style = Font(bold=True) if bold_bool else Font()
        style_range(ws, 'C' + num, letter + num, font=bold_style, border=border_style)

    if sheet_name == IS:
        style_row(ws, "total sales", IS, is_df, bs_df, cf_df, False)
        style_row(ws, "gross income", IS, is_df, bs_df, cf_df)
        style_row(ws, "ebit operating income", IS, is_df, bs_df, cf_df)
        style_row(ws, "ebit operating income", IS, is_df, bs_df, cf_df)
        style_row(ws, "pretax income", IS, is_df, bs_df, cf_df)
        style_row(ws, "net income", IS, is_df, bs_df, cf_df)
        style_row(ws, "driver ratio", IS, is_df, bs_df, cf_df, underline="single",
                  border_bool=False)
    elif sheet_name == BS:
        style_row(ws, "total shareholder equity", BS, is_df, bs_df, cf_df, bold_bool=False,
                  border_bool=False)
        style_row(ws, "total liabilit shareholder equity", BS, is_df, bs_df, cf_df,
                  border_bool=False)
        style_row(ws, "driver ratio", BS, is_df, bs_df, cf_df, underline="single",
                  border_bool=False)
    elif sheet_name == CF:
        style_row(ws, "net operating cash flow cf", CF, is_df, bs_df, cf_df)
        style_row(ws, "cash balance", CF, is_df, bs_df, cf_df, border_bool=False)
        style_row(ws, "driver ratio", CF, is_df, bs_df, cf_df, underline="single",
                  border_bool=False)

def main():
    income_statement = pd.read_excel("NVIDIA/NVIDIA Income Statement.xlsx", header=4,
                                     index_col=0)
    balance_sheet = pd.read_excel("NVIDIA/NVIDIA Balance Sheet.xlsx", header=4, index_col=0)
    cash_flow = pd.read_excel("NVIDIA/NVIDIA Cash Flow.xlsx", header=4, index_col=0)

    income_statement, _ = preprocess(income_statement)
    balance_sheet, _ = preprocess(balance_sheet)
    cash_flow, fye = preprocess(cash_flow)

    # FIXME temporary slices of data
    income_statement = income_statement[:21]
    balance_sheet = balance_sheet[:31]
    cash_flow = cash_flow[:26]

    # FIXME temporary filter in labels
    filter_in(income_statement, [
        "sales", "cost of good sold", "gross income", "sg&a expense", "other expense",
        "ebit operating income", "nonoperating income net", "interest expense",
        "unusual expense", "income taxes", "dilute eps", "net income", "div per share", "ebitda"
    ])

    # FIXME temporary parameters
    growth_rates = [0.5, 0.5, 0.5, 0.5, 0.5]
    yrs_to_predict = len(growth_rates)

    # Cast year data type
    income_statement.columns = income_statement.columns.astype(int)
    balance_sheet.columns = balance_sheet.columns.astype(int)
    cash_flow.columns = cash_flow.columns.astype(int)

    income_statement, cash_flow = process_is(income_statement, cash_flow, growth_rates,
                                             yrs_to_predict)
    balance_sheet, cash_flow = process_bs(income_statement, balance_sheet, cash_flow,
                                          yrs_to_predict)
    cash_flow = process_cf(income_statement, balance_sheet, cash_flow, yrs_to_predict)

    wb = openpyxl.Workbook()
    wb['Sheet'].title = IS
    wb.create_sheet(BS)
    wb.create_sheet(CF)
    for r in dataframe_to_rows(income_statement):
        wb[IS].append(r)
    for r in dataframe_to_rows(balance_sheet):
        wb[BS].append(r)
    for r in dataframe_to_rows(cash_flow):
        wb[CF].append(r)
    style_ws(wb[IS], IS, income_statement, balance_sheet, cash_flow, fye)
    style_ws(wb[BS], BS, income_statement, balance_sheet, cash_flow, fye)
    style_ws(wb[CF], CF, income_statement, balance_sheet, cash_flow, fye)
    wb.save("output.xlsx")

if __name__ == "__main__":
    main()
